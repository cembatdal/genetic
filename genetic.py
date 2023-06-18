import random, time, openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

df = pd.read_excel('150jobs.xlsx', sheet_name='masterData')

df['ProcessTime'] = round((df['Miktar_Mt'] / df['Hiz']), 2)

df = df.reset_index(drop=True)

df.index = df.index + 1

df = df.rename_axis('Index')

master_data = df[['IsNo', 'Recete', 'Sicaklik', 'ProcessTime']]

master_data.set_index("IsNo", inplace=True)

is_number = master_data.index.tolist()

total_process_time = master_data['ProcessTime'].sum()

#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@

dataframe_recipe_names = pd.read_excel('150jobs.xlsx', sheet_name='recipeMatrix')

recipe_names = dataframe_recipe_names.columns.tolist()

recipe_names.pop(0)

dataframe_recipe_matrix = pd.read_excel('150jobs.xlsx', sheet_name='recipeMatrix')

values_recipe_matrix = dataframe_recipe_matrix.values

recipe_matrix = values_recipe_matrix.tolist()

recipe_matrix = [satir[1:] for satir in recipe_matrix]

dataframe = pd.DataFrame(recipe_matrix, columns=recipe_names, index=recipe_names)

#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
#Farklı Sıcaklıkları Dinamik Olarak Tespit Etme ve İş Numaralarını Gruplandırma
#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@

different_temperatures = sorted(master_data['Sicaklik'].unique())
print(different_temperatures)

temperature_lists = {}
for temperature in different_temperatures:
    filtre = master_data['Sicaklik'] == temperature
    work_orders = master_data.loc[filtre].index.tolist()
    temperature_lists[temperature] = work_orders

#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
#Fonksiyonlar

def get_recipe(isno):
    return master_data.loc[isno, "Recete"]

def get_prep_time(recipe_1, recipe_2):
    return dataframe.loc[recipe_1, recipe_2]

def get_temperature(isno):
    return master_data.loc[isno, "Sicaklik"]

def get_temperature_difference(temp_1, temp_2):
    return temp_1 <= temp_2

def create_population(population_size, chromosome_size, isnolist):
    population = []
    for i in range(population_size):
        chromosome = random.sample(isnolist, chromosome_size)
        population.append(chromosome)
    return population

def fitness(chromosome):
    total_prep_time = 0
    for i in range(len(chromosome) - 1):
        current_job = chromosome[i]
        next_job = chromosome[i+1]
        total_prep_time += get_prep_time(get_recipe(current_job),get_recipe(next_job))
    return total_prep_time

def calculate_total_prep_time(chromosome):
    total_prep_time = 0
    for i in range(len(chromosome) - 1):
        current_job = chromosome[i]
        next_job = chromosome[i+1]
        total_prep_time += get_prep_time(get_recipe(current_job),get_recipe(next_job))
    return total_prep_time

def fitness_list(chromosome):
    fitness_list = []
    for i in range(len(chromosome) - 1):
        current_job = chromosome[i]
        next_job = chromosome[i+1]
        fitness_list.append(get_prep_time(get_recipe(current_job),get_recipe(next_job)))
    fitness_list.append(0)
    return fitness_list

def selection(population):
    parent1 = random.choice(population)
    parent2 = random.choice(population)
    return parent1, parent2

def pmx_crossover(parent1, parent2, part_length):
    n = len(parent1)
    while True:
        crossover_point1 = random.randint(0, n-1)
        crossover_point2 = random.randint(0, n-1)
        while crossover_point2 == crossover_point1:
            crossover_point2 = random.randint(0, n-1)
        if abs(crossover_point2 - crossover_point1) == part_length:
            break 
    if crossover_point1 > crossover_point2:
        crossover_point1, crossover_point2 = crossover_point2, crossover_point1

    child1 = [-1] * n
    child2 = [-1] * n

    for i in range(crossover_point1, crossover_point2+1):
        child1[i] = parent1[i]

    for i in range(crossover_point1, crossover_point2+1):
        if parent2[i] not in child1:
            j = i
            while parent1[j] in child1[crossover_point1:crossover_point2+1]:
                j = parent2.index(parent1[j])
            child1[j] = parent2[i]

    for i in range(n):
        if child1[i] == -1:
            child1[i] = parent2[i]

    for i in range(crossover_point1, crossover_point2+1):
        child2[i] = parent2[i]

    for i in range(crossover_point1, crossover_point2+1):
        if parent1[i] not in child2:
            j = i
            while parent2[j] in child2[crossover_point1:crossover_point2+1]:
                j = parent1.index(parent2[j])
            child2[j] = parent1[i]

    for i in range(n):
        if child2[i] == -1:
            child2[i] = parent1[i]
    return child1, child2

def mutation(chromosome, mutation_rate):
    if random.random() < mutation_rate:
        i, j = random.sample(range(len(chromosome)), 2)
        chromosome[i], chromosome[j] = chromosome[j], chromosome[i]
    return chromosome

def genetic_algorithm(population_size, chromosome_size, generations, mutation_rate, part_length, isnolist):
    population = create_population(population_size, chromosome_size, isnolist)

    for generation in range(generations):
        new_population = []

        population = sorted(population, key=lambda chromosome: fitness(chromosome), reverse=False)
        print("Nesil:", generation, "\tAra uygunluk değeri:", fitness(population[0]), )
        new_population.append(population[0])

        while len(new_population) < population_size:
            parent1, parent2 = selection(population)
            child1, child2 = pmx_crossover(parent1, parent2, part_length)
            child1 = mutation(child1, mutation_rate)
            child2 = mutation(child2, mutation_rate)
            new_population.append(child1)
            if len(new_population) < population_size:
                new_population.append(child2)

        population = new_population

    population = sorted(population, key=lambda chromosome: fitness(chromosome), reverse=False)
    return population[0]

##############################################################
#Runtime Parameters and Main Function
##############################################################

isnolist = is_number

population_size = 50

generations = 50

mutation_rate = 0.2

start = time.time()

genetic_lists = {}
for temperature, work_orders in temperature_lists.items():
    chromosome_size = len(work_orders)
    isnolist = work_orders
    part_length = round((chromosome_size / 2), 0)
    genetic_lists[temperature] = genetic_algorithm(population_size=population_size, chromosome_size=chromosome_size, generations=generations, mutation_rate=mutation_rate, part_length=part_length, isnolist=work_orders)

merged_list = []
for list in genetic_lists.values():
    merged_list.extend(list)

result = merged_list

##############################################################
#Output
##############################################################

print("-"*150)

print(chromosome_size, "iş için en iyi sıralama:\t", result)

print("-"*150)

fitness_value = fitness(result)

total_prep_time = calculate_total_prep_time(result)

print("Uygunluk değeri:\t\t", fitness_value, "\tdakika")

print("Reçeteler arası geçiş süresi:\t", total_prep_time, "\tdakika")

print("Kumaş işlem süresi:\t\t", round(total_process_time), "\tdakika")

print("Toplam işlem süresi:\t\t", round((fitness_value + total_process_time)), "\tdakika")

print("-"*150)

end = time.time()

algorithm_run_time = round(end-start, 2)

print("Algoritmanın çalışma süresi:\t", algorithm_run_time, "\tsaniye")

print("-"*150)

##############################################################
#Write experiment data to .xlsx file
##############################################################

def find_next_empty_row(sheet):
    row_num = 1
    while sheet.cell(row=row_num, column=1).value is not None:
        row_num += 1
    return row_num

def append_dataframe_to_excel(df, file_path, sheet_name):
    # Excel dosyasını yükle
    workbook = openpyxl.load_workbook(file_path)

    # Hedef sayfayı seç veya oluştur
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        next_row = sheet.max_row + 1
    else:
        sheet = workbook.create_sheet(sheet_name)
        next_row = 1

    # DataFrame'i satır satır kopyala ve sayfaya ekle
    for row in dataframe_to_rows(df, index=False, header=False):
        for col_num, value in enumerate(row, start=1):
            sheet.cell(row=next_row, column=col_num, value=value)
        next_row += 1

    # Dosyayı kaydet
    workbook.save(file_path)

dosya_yolu = '150_jobs_parameter_experiments.xlsx'

veriler = {
    'pop_size': [population_size], 
    'chromosome_size': [chromosome_size], 
    'generations': [generations],
    'mutation_rate': [mutation_rate],
    'part_length': [part_length],
    'algorithm_run_time': [algorithm_run_time],
    'fitness_value': [fitness_value]
    }

df = pd.DataFrame(veriler)
sayfa_adi = 'generations'

append_dataframe_to_excel(df, dosya_yolu, sayfa_adi)

print(result)
